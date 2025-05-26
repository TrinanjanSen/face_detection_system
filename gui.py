from tkinter import *
from tkinter import messagebox
from PIL import Image, ImageTk
from face_attendance_system import capture_face_images, recognize_and_mark_attendance
from openpyxl import load_workbook
from collections import defaultdict
import os

def capture_face():
    name = name_entry.get()
    if not name:
        messagebox.showwarning("Input Error", "Please enter a name.")
        return
    capture_face_images(name)  
    messagebox.showinfo("Success", f"Images for {name} captured.")

def start_attendance():
    try:
        gap_minutes = int(attendance_gap_entry.get())
        if gap_minutes <= 0:
            raise ValueError
    except:
        messagebox.showwarning("Input Error", "Invalid attendance gap. Using default 30 minutes.")
        gap_minutes = 30

    recognize_and_mark_attendance(timeout=60, attendance_gap_minutes=gap_minutes)
    messagebox.showinfo("Done", f"Attendance marking completed with a gap of {gap_minutes} minutes.")

def show_attendance_percentage():
    file_name = 'attendance.xlsx'
    if not os.path.exists(file_name):
        messagebox.showwarning("File Error", "Attendance file not found.")
        return

    wb = load_workbook(file_name)
    ws = wb.active

    attendance = defaultdict(lambda: {'P':0, 'A':0})
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, datetime_str, status = row
        if status == 'P':
            attendance[name]['P'] +=1
        elif status == 'A':
            attendance[name]['A'] +=1

    message = ""
    for name, counts in attendance.items():
        total = counts['P'] + counts['A']
        percentage = (counts['P'] / total) * 100 if total else 0
        message += f"{name}: {percentage:.2f}% ({counts['P']} P, {counts['A']} A)\n"

    messagebox.showinfo("Attendance %", message)

def exit_app():
    root.destroy()

root = Tk()
root.title("Face Attendance")
root.geometry("700x400")
root.configure(bg="#2E4053")

header = Label(root, text="Face Recognition Attendance System", font=("Helvetica", 20, "bold"), fg="white", bg="#2E4053")
header.pack(pady=20)

name_frame = Frame(root, bg="#2E4053")
name_frame.pack(pady=10)

Label(name_frame, text="Enter Name:", font=("Helvetica", 14), fg="white", bg="#2E4053").pack(side="left", padx=5)
name_entry = Entry(name_frame, font=("Helvetica", 14), width=20)
name_entry.pack(side="left", padx=5)

window_frame = Frame(root, bg="#2E4053")
window_frame.pack(pady=10)

Label(window_frame, text="Attendance Gap (minutes):", font=("Helvetica", 14), fg="white", bg="#2E4053").pack(side="left", padx=5)
attendance_gap_entry = Entry(window_frame, font=("Helvetica", 14), width=5)
attendance_gap_entry.pack(side="left", padx=5)
attendance_gap_entry.insert(0, "30")

icon_size = (64, 64)
capture_icon = ImageTk.PhotoImage(Image.open("icons/capture.png").resize(icon_size))
attendance_icon = ImageTk.PhotoImage(Image.open("icons/attendance.png").resize(icon_size))
exit_icon = ImageTk.PhotoImage(Image.open("icons/exit.png").resize(icon_size))

button_frame = Frame(root, bg="#2E4053")
button_frame.pack(pady=20)

Button(button_frame, image=capture_icon, text="Capture Face", compound="top", font=("Helvetica", 12), fg="white", bg="#1ABC9C", command=capture_face, width=120, height=120).grid(row=0, column=0, padx=20)
Button(button_frame, image=attendance_icon, text="Start Attendance", compound="top", font=("Helvetica", 12), fg="white", bg="#3498DB", command=start_attendance, width=120, height=120).grid(row=0, column=1, padx=20)
Button(button_frame, text="Attendance %", font=("Helvetica", 12), fg="white", bg="#F39C12", command=show_attendance_percentage, width=15, height=2).grid(row=0, column=2, padx=20)
Button(button_frame, image=exit_icon, text="Exit", compound="top", font=("Helvetica", 12), fg="white", bg="#E74C3C", command=exit_app, width=120, height=120).grid(row=0, column=3, padx=20)

root.mainloop()
