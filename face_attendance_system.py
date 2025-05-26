import cv2
import face_recognition
import numpy as np
import os
from datetime import datetime, timedelta
import time
from openpyxl import Workbook, load_workbook

def capture_face_images(person_name, num_images=12):
    base_folder = 'images'
    person_folder = os.path.join(base_folder, person_name)
    os.makedirs(person_folder, exist_ok=True)

    cap = cv2.VideoCapture(0)
    print(f"Press SPACE to capture {num_images} images for {person_name}, ESC to cancel.")

    captured_count = 0

    while captured_count < num_images:
        ret, frame = cap.read()
        if not ret:
            break

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)

        for (top, right, bottom, left) in face_locations:
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

        cv2.imshow("Capture Face", frame)
        key = cv2.waitKey(1)

        if key % 256 == 27:
            print("Capture cancelled.")
            break
        elif key % 256 == 32:
            if face_locations:
                img_path = os.path.join(person_folder, f"{person_name}_{captured_count+1}.jpg")
                cv2.imwrite(img_path, frame)
                print(f"Captured image {captured_count+1} at {img_path}")
                captured_count += 1
                time.sleep(0.5)
            else:
                print("No face detected! Try again.")

    cap.release()
    cv2.destroyAllWindows()

def load_known_faces(known_faces_dir='images'):
    known_encodings = []
    known_names = []

    for person_name in os.listdir(known_faces_dir):
        person_folder = os.path.join(known_faces_dir, person_name)
        if os.path.isdir(person_folder):
            for filename in os.listdir(person_folder):
                if filename.lower().endswith((".jpg", ".png")):
                    img_path = os.path.join(person_folder, filename)
                    image = face_recognition.load_image_file(img_path)
                    encodings = face_recognition.face_encodings(image)
                    if encodings:
                        known_encodings.append(encodings[0])
                        known_names.append(person_name)
                    else:
                        print(f"Warning: No face found in {filename}")
    return known_encodings, known_names

def mark_attendance_excel(name, attendance_gap_minutes=30):
    file_name = 'attendance.xlsx'
    now = datetime.now()
    now_str = now.strftime('%Y-%m-%d %H:%M:%S')

    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'DateTime', 'Status'])
        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active

    already_marked = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == name:
            row_datetime = datetime.strptime(row[1], '%Y-%m-%d %H:%M:%S')
            if (now - row_datetime) < timedelta(minutes=attendance_gap_minutes):
                already_marked = True
                break

    if not already_marked:
        ws.append([name, now_str, 'P'])
        wb.save(file_name)
        print(f"Attendance marked for {name} at {now_str}")
        return True
    else:
        print(f"Attendance for {name} already marked within {attendance_gap_minutes} minutes.")
        return False

def recognize_and_mark_attendance(timeout=120, attendance_gap_minutes=30):
    known_encodings, known_names = load_known_faces()
    if not known_encodings:
        print("No known faces found. Please add images first.")
        return "No known faces found."

    cap = cv2.VideoCapture(0)
    print("Starting Face Recognition Attendance System...")

    start_time = time.time()
    marked_names = set()
    message_log = []

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        for face_encoding, face_location in zip(face_encodings, face_locations):
            matches = face_recognition.compare_faces(known_encodings, face_encoding)
            name = "Unknown"

            face_distances = face_recognition.face_distance(known_encodings, face_encoding)
            if len(face_distances) > 0:
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    name = known_names[best_match_index]
                    if name not in marked_names:
                        marked = mark_attendance_excel(name, attendance_gap_minutes)
                        message = (f"Attendance marked for {name}" if marked else
                                   f"Attendance for {name} already marked within {attendance_gap_minutes} minutes")
                        print(message)
                        message_log.append(message)
                        marked_names.add(name)

            top, right, bottom, left = [v * 4 for v in face_location]
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            if name in marked_names:
                cv2.putText(frame, "Attendance processed", (left, bottom + 30),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 0), 2)

        cv2.imshow('Attendance System', frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            print("User pressed 'q'. Exiting...")
            break
        if time.time() - start_time > timeout:
            print(f"Timeout reached ({timeout}s), closing...")
            break

    cap.release()
    cv2.destroyAllWindows()

    return "\n".join(message_log) if message_log else "No attendance marked."

def cli_main():
    while True:
        print("\n=== Face Recognition Attendance System ===")
        print("1. Capture New Face Images")
        print("2. Start Attendance System")
        print("3. Exit")
        choice = input("Enter your choice (1/2/3): ")

        if choice == '1':
            person_name = input("Enter the name of the person to capture: ")
            capture_face_images(person_name)
        elif choice == '2':
            recognize_and_mark_attendance()
        elif choice == '3':
            print("Exiting...")
            break
        else:
            print("Invalid choice. Try again.")

if __name__ == "__main__":
    cli_main()
